from rest_framework import viewsets, status
from rest_framework.response import Response
from rest_framework.decorators import action,api_view, permission_classes
from rest_framework.parsers import FileUploadParser
from rest_framework.permissions import IsAuthenticated,AllowAny
from rest_framework_simplejwt.authentication import JWTAuthentication
from django.contrib.auth.models import User
from django.shortcuts import render
from django.http import HttpResponse
from django.views.generic import View
from django.contrib.auth.decorators import login_required
from .models import *
from .seriealizers import *
from rest_framework_simplejwt.views import TokenObtainPairView
from rest_framework.permissions import AllowAny, IsAuthenticated
from .resources import ClientResource
from openpyxl import load_workbook
import csv


class JWTAuthenticationMixin:
    authentication_classes = [JWTAuthentication]
    permission_classes = [IsAuthenticated]

class ClientViewSet(JWTAuthenticationMixin, viewsets.ModelViewSet):
    queryset = Clients.objects.all()
    serializer_class = ClientSerializer

class BillViewSet(JWTAuthenticationMixin, viewsets.ModelViewSet):
    queryset = Bills.objects.all()
    serializer_class = BillSerializer

class ProductViewSet(JWTAuthenticationMixin, viewsets.ModelViewSet):
    queryset = Products.objects.all()
    serializer_class = ProductSerializer

class UserViewSet(viewsets.ModelViewSet):
    queryset = User.objects.all()
    serializer_class = UserSerializer

    @action(detail=False, methods=['post'])
    def register_user(self, request):
        serializer = self.serializer_class(data=request.data)
        if serializer.is_valid():
            user = serializer.save()
            return Response({'status': 'Usuario registrado con éxito'}, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

class MyTokenObtainPairView(TokenObtainPairView):
    permission_classes = [AllowAny]

token_obtain_pair = MyTokenObtainPairView.as_view()

def export_clients_csv(request):
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = "attachment; filename='clients.csv'"

    # Obtener todos los clientes con la cantidad de facturas relacionadas
    clients = Clients.objects.annotate(num_bills=models.Count('bills'))

    # Crear el objeto de escritura CSV
    writer = csv.writer(response)
    
    # Escribir la cabecera del CSV
    writer.writerow(['Documento', 'Nombre Completo', 'Cantidad de Facturas'])

    # Escribir los datos de cada cliente
    for client in clients:
        writer.writerow([client.document, client.full_name(), client.num_bills])

    return response

from django.http import JsonResponse


@api_view(['POST'])
@permission_classes([AllowAny])
def import_from_excel(request):
    if request.method == 'POST':
        excel_file = request.FILES['pruebas_api']

        # Comprueba si el archivo es un archivo Excel
        if not excel_file.name.endswith('.xlsx'):
            return JsonResponse({'error': 'El archivo no es de tipo XLSX.'}, status=400)

        try:
            wb = load_workbook(excel_file)
            ws = wb.active

            success_count = 0
            error_count = 0
            errors = []

            for row in ws.iter_rows(min_row=2, values_only=True):
                document, first_name, last_name, email, password = row
                try:
                    Clients.objects.create(
                        document=document,
                        first_name=first_name,
                        last_name=last_name,
                        email=email,
                        password=password
                    )
                    success_count += 1
                except Exception as e:
                    error_count += 1
                    errors.append({'row_data': row, 'error_message': str(e)})

            response_data = {
                'success_count': success_count,
                'error_count': error_count,
                'errors': errors
            }

            return JsonResponse(response_data)
        except Exception as e:
            return JsonResponse({'error': f'Error durante la importación: {str(e)}'}, status=500)

    return JsonResponse({'error': 'Método no permitido.'}, status=405)

